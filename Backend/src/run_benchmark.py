"""
Benchmark runner — reads questions from benchmark.xlsx, runs each through
qa_pipeline, and writes SQL + NL answer + response time + steward metrics
(``sql_agent_llm_rounds``, ``sql_agent_sql_steps``, ``error``) back into the sheet.

Run from Backend/src:
    python run_benchmark.py

Close benchmark.xlsx in Excel before running (file is locked while open).
Optional: set BENCHMARK_FILE env var to a different path.
"""

from __future__ import annotations

import os
import sys
import time
import traceback
from pathlib import Path

# ── project root on sys.path ────────────────────────────────────────────────
_SRC = Path(__file__).resolve().parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

# ── load .env before any project imports ────────────────────────────────────
from env_loader import (
    force_apply_azure_openai_from_dotenv,
    force_apply_redis_from_dotenv,
    load_application_dotenv,
)

load_application_dotenv()
force_apply_azure_openai_from_dotenv()
force_apply_redis_from_dotenv()

# ── load workbook into SQLite (same as api_server lifespan) ─────────────────
from db_adapter import use_sqlite_backend

if use_sqlite_backend():
    from config import settings
    from data_loader import get_db, load_file

    print(f"[benchmark] Loading workbook: {settings.data_file_path}", flush=True)
    load_file(settings.data_file_path)
    db0 = get_db()
    if db0 is None:
        print("[benchmark] ERROR: workbook did not load — check DATA_FILE_PATH in .env", file=sys.stderr)
        sys.exit(2)
    print(f"[benchmark] Workbook ready: {db0.file_name} ({len(db0.tables)} tables)", flush=True)
    try:
        from workbook_schema_rag import ensure_workbook_rag_index
        ensure_workbook_rag_index(db0)
    except Exception as rag_exc:
        print(f"[benchmark] workbook RAG skipped: {rag_exc}", file=sys.stderr)

# ── pipeline imports ─────────────────────────────────────────────────────────
from conversation_context import ConversationBuffer
from qa_pipeline import run_question_pipeline_turn

# ── openpyxl ─────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font
except ImportError:
    print("[benchmark] openpyxl not found — pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── resolve benchmark file path ──────────────────────────────────────────────
BENCHMARK_FILE = Path(os.getenv("BENCHMARK_FILE", str(_SRC / "benchmark.xlsx"))).resolve()

if not BENCHMARK_FILE.exists():
    print(f"[benchmark] File not found: {BENCHMARK_FILE}", file=sys.stderr)
    sys.exit(1)

print(f"[benchmark] Reading: {BENCHMARK_FILE}", flush=True)

# ── open workbook ─────────────────────────────────────────────────────────────
try:
    wb = openpyxl.load_workbook(BENCHMARK_FILE)
except PermissionError:
    print(
        "[benchmark] PermissionError — close benchmark.xlsx in Excel first, then retry.",
        file=sys.stderr,
    )
    sys.exit(1)

ws = wb.active

# ── detect column indices from header row ─────────────────────────────────────
header = {str(ws.cell(1, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}

def _col(candidates: list[str]) -> int | None:
    for name in candidates:
        if name in header:
            return header[name]
    return None

COL_Q   = _col(["questions", "question", "q"])
COL_SQL = _col(["sql", "generated_sql", "sql_query"])
COL_GEN = _col(["generated", "nl_answers", "nl_answer", "answer", "response", "generated_answer"])
COL_RT  = _col(["response_time", "response time", "time_ms", "latency_ms", "time"])
COL_LLM = _col(["sql_agent_llm_rounds", "llm_rounds", "steward_llm_rounds", "azure_rounds"])
COL_STP = _col(["sql_agent_sql_steps", "sql_steps", "steward_sql_steps"])
COL_ERR = _col(["error", "pipeline_error", "err"])

if COL_Q is None:
    print(f"[benchmark] Could not find a 'questions' column. Found headers: {list(header.keys())}", file=sys.stderr)
    sys.exit(1)

# Add missing columns
def _ensure_col(col_idx: int | None, label: str) -> int:
    if col_idx is not None:
        return col_idx
    next_col = ws.max_column + 1
    ws.cell(1, next_col).value = label
    ws.cell(1, next_col).font = Font(bold=True)
    return next_col

COL_SQL = _ensure_col(COL_SQL, "sql")
COL_GEN = _ensure_col(COL_GEN, "nl_answers")
COL_RT  = _ensure_col(COL_RT,  "response_time")
COL_LLM = _ensure_col(COL_LLM, "sql_agent_llm_rounds")
COL_STP = _ensure_col(COL_STP, "sql_agent_sql_steps")
COL_ERR = _ensure_col(COL_ERR, "error")

print(
    f"[benchmark] Columns → Q:{COL_Q}  SQL:{COL_SQL}  NL:{COL_GEN}  RT:{COL_RT}  "
    f"LLM:{COL_LLM}  SQLsteps:{COL_STP}  ERR:{COL_ERR}",
    flush=True,
)

# ── highlight styles ──────────────────────────────────────────────────────────
FILL_OK  = PatternFill("solid", fgColor="C6EFCE")  # green
FILL_ERR = PatternFill("solid", fgColor="FFC7CE")  # red

# ── iterate rows ──────────────────────────────────────────────────────────────
total = 0
for row in range(2, ws.max_row + 1):
    question = ws.cell(row, COL_Q).value
    if not question or not str(question).strip():
        continue

    question = str(question).strip()
    total += 1
    print(f"\n[{total}] Q: {question[:100]}{'…' if len(question) > 100 else ''}", flush=True)

    buf = ConversationBuffer()
    t_start = time.perf_counter()
    llm_rounds: int | str = ""
    sql_steps: int | str = ""
    err_val = ""

    try:
        out = run_question_pipeline_turn(question, conversation=buf, use_cache=False)
        elapsed_ms = int((time.perf_counter() - t_start) * 1000)

        sql_val = out.get("sql") or ""
        ans_val = out.get("answer") or ""
        err_val = (out.get("error") or "").strip()
        if not ans_val and err_val:
            ans_val = err_val
        ok = bool(out.get("answer"))

        lr = out.get("sql_agent_llm_rounds")
        ss = out.get("sql_agent_sql_steps")
        llm_rounds = int(lr) if lr is not None and str(lr).strip() != "" else ""
        sql_steps = int(ss) if ss is not None and str(ss).strip() != "" else ""

    except Exception:
        elapsed_ms = int((time.perf_counter() - t_start) * 1000)
        sql_val = ""
        ans_val = ""
        err_val = traceback.format_exc(limit=5)
        ok = False

    # write back
    def _write(col: int, value: str) -> None:
        cell = ws.cell(row, col)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = FILL_OK if ok else FILL_ERR

    _write(COL_SQL, sql_val)
    _write(COL_GEN, ans_val if ans_val else (err_val or ""))
    ws.cell(row, COL_RT).value = elapsed_ms
    ws.cell(row, COL_RT).alignment = Alignment(vertical="top")

    ws.cell(row, COL_LLM).value = llm_rounds
    ws.cell(row, COL_LLM).alignment = Alignment(vertical="top")
    ws.cell(row, COL_LLM).fill = FILL_OK if ok else FILL_ERR

    ws.cell(row, COL_STP).value = sql_steps
    ws.cell(row, COL_STP).alignment = Alignment(vertical="top")
    ws.cell(row, COL_STP).fill = FILL_OK if ok else FILL_ERR

    _write(COL_ERR, err_val)

    print(
        f"    → {'OK' if ok else 'ERR'}  {elapsed_ms}ms  llm_rounds={llm_rounds!s}  sql_steps={sql_steps!s}",
        flush=True,
    )
    if sql_val:
        print(f"    SQL: {sql_val[:120]}{'…' if len(sql_val) > 120 else ''}", flush=True)

    # save after every row so partial results are not lost on crash/interrupt
    try:
        wb.save(BENCHMARK_FILE)
    except PermissionError:
        out_path = BENCHMARK_FILE.with_name(BENCHMARK_FILE.stem + "_results.xlsx")
        wb.save(out_path)
        print(f"    [benchmark] benchmark.xlsx locked — saved partial results to {out_path}", file=sys.stderr)

print(f"\n[benchmark] Done — {total} questions processed. Results saved to {BENCHMARK_FILE}", flush=True)
